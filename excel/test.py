from openpyxl import load_workbook
from openpyxl import Workbook
from itertools import combinations

wb = load_workbook("v1.xlsx")

if len(wb.sheetnames) > 0:

    newWb = Workbook()
    newSheet = newWb.active

    sheet = wb[wb.sheetnames[0]]

    for row in sheet.rows:
        rowList = list()

        for column in row:
            if len(column.value or ()) != 0:
                rowList.append(column.value)

        for c in combinations(rowList, 2):
            newSheet.append(c)

    newWb.save("new.xlsx")
