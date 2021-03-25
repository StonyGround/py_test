from openpyxl import load_workbook
from openpyxl import Workbook
from itertools import combinations

import os

for file in os.listdir("./"):
    if file.endswith(".xlsx"):
        print("读取到文件[" + file + "]")
        wb = load_workbook(file)

        if len(wb.sheetnames) > 0:

            newWb = Workbook()
            newSheet = newWb.active

            sheet = wb[wb.sheetnames[0]]
            print("正在读取文件...")
            for row in sheet.rows:
                rowList = list()

                for column in row:
                    if len(column.value or ()) != 0:
                        rowList.append(column.value)

                for c in combinations(rowList, 2):
                    newSheet.append(c)
            print("写入完成，正在保存...")
            newWb.save("new-" + file)
        else:
            print("未读取到sheet表格")
