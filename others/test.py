from openpyxl import load_workbook
from selenium import webdriver

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
# 登录
driver.get("http://daily-s.beicaizs.com/web/index.html#/account/login")
driver.find_element_by_id("accountName").send_keys("北京天相投顾有限公司")
driver.find_element_by_id("accountPassword").send_keys("Aa123456")
driver.find_element_by_class_name("ant-btn").click()
# 点击合规系统
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="layoutMain"]/div[2]/div[2]/div/div/div/div'))).click()
# 系统设置
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/section/aside/div/div/div[1]/div/ul/li[3]/div[1]'))).click()
# 角色管理
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/section/aside/div/div/div[1]/div/ul/li[3]/ul/li[1]'))).click()

# wb = load_workbook('1.xlsx')
# sheet = wb[wb.sheetnames[0]]
# for row in sheet['F':'AA']:
#   for column in row:
#     print(column.value)

# 增加按钮
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="layoutMain"]/div[2]/div[2]/div/div/div[1]/form/div/div[4]/button[3]'))).click()
